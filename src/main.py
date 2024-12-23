import time
import warnings
import chromedriver_autoinstaller
import requests
import urllib3
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from datetime import datetime, timedelta
import os
import sys
import configparser
import json
from tqdm import tqdm
import openpyxl
from datetime import datetime, timezone
import subprocess


# Ignorar warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
urllib3.disable_warnings()
requests.packages.urllib3.disable_warnings()
import sys




LOCAL_VERSION = "1.0.0"
VERSION_JSON_URL = "https://raw.githubusercontent.com/Raos48/ProtocolarTarefasPat/main/updates/version.json"

def main():

    def check_for_update():
        """
        Faz requisição ao version.json no GitHub,
        compara com a versão local e, se houver versão nova,
        pergunta ao usuário se deseja atualizar.
        """
        print(f"Sua versão atual: {LOCAL_VERSION}")
        try:
            response = requests.get(VERSION_JSON_URL, timeout=10)
            if response.status_code == 200:
                data = response.json()
                latest_version = data.get("version")
                file_url = data.get("file_url")
                if latest_version and file_url:
                    if latest_version > LOCAL_VERSION:
                        print(f"Há uma nova versão disponível: {latest_version}")
                        choice = input("Deseja atualizar agora? (S/N): ")
                        if choice.upper() == "S":
                            download_and_replace(file_url, latest_version)
                    else:
                        print("Você já está na versão mais recente.")
                else:
                    print("Erro: JSON de versão não contém os campos esperados.")
            else:
                print(f"Não foi possível verificar updates. HTTP {response.status_code}")
        except Exception as e:
            print("Erro ao verificar updates:", e)


    def download_and_replace(url, new_version):
        """
        Faz o download do novo executável e substitui o antigo.
        Em seguida, relança a nova versão, se desejar.
        """
        exe_name = f"main_v{new_version}.exe"
        temp_exe = f"update_{new_version}.exe"

        try:
            print("Baixando nova versão...")
            r = requests.get(url, stream=True)
            if r.status_code == 200:
                with open(temp_exe, "wb") as f:
                    for chunk in r.iter_content(chunk_size=8192):
                        f.write(chunk)
                print("Download concluído.")

                # Fechar a aplicação antiga (opcional)
                # Se você estiver rodando como script .py, talvez não precise fechar,
                # mas se estiver em .exe, não pode substituir a si mesmo em uso.

                # Exemplo: renomear o executável atual (main.exe) para "main_old.exe"
                # e depois renomear "update_new.exe" -> "main.exe"
                old_exe = "main.exe"  # Nome do executável atual
                if os.path.exists(old_exe):
                    os.rename(old_exe, f"main_old_{LOCAL_VERSION}.exe")

                # Renomear o baixado para main.exe
                os.rename(temp_exe, "main.exe")
                print("Aplicação atualizada para a versão", new_version)

                # (Opcional) relançar nova versão e sair
                print("Iniciando nova versão...")
                subprocess.Popen(["main.exe"])
                sys.exit(0)  # Fecha o atual

            else:
                print("Falha no download da nova versão. HTTP status:", r.status_code)
        except Exception as e:
            print("Erro no processo de atualização:", e)

    check_for_update()

    def find_excel_file(filename):
        possible_paths = [
            os.path.join(os.path.dirname(sys.executable), filename),
            os.path.join(os.getcwd(), filename)
        ]

        for path in possible_paths:
            if os.path.exists(path):
                return path
        
        return None


    excel_filename = "protocolo_pat.xlsx"
    excel_path = find_excel_file(excel_filename)


    if excel_path:
        print(f"Arquivo Excel encontrado em: {excel_path}")
        try:
            df = pd.read_excel(excel_path)
            print("Arquivo Excel carregado com sucesso.")
        except Exception as e:
            print(f"Erro ao carregar o arquivo Excel: {str(e)}")
            input("Pressione Enter para sair...")
            sys.exit(1)
    else:
        print(f"Erro: O arquivo Excel '{excel_filename}' não foi encontrado.")
        print("Certifique-se de que o arquivo está no mesmo diretório que o executável.")
        input("Pressione Enter para sair...")
        sys.exit(1)

    print("Verificando versão Chromedriver...")
    chromedriver_autoinstaller.install()
    print("Checagem Chromedriver Finalizada...")






    def obter_token():
        print("Obter Token..")
        warnings.filterwarnings("ignore", category=DeprecationWarning)
        urllib3.disable_warnings()
        requests.packages.urllib3.disable_warnings()
        chromedriver_autoinstaller.install()
        
        
        config = configparser.ConfigParser()
        config.read('config.ini')
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        

        
        
        # Carregar headers do arquivo
        headers_file_path = os.path.join(os.getcwd(), config['FILES']['headers_file_path'])
        try:
            with open(headers_file_path, 'r') as file:
                headers = json.loads(file.read())
            print("Headers carregados com sucesso.")
        except Exception as e:
            print(f"Erro ao carregar headers: {e}")
            headers = {}

        # Fazer requisição à API
        url = config['API_URLS']['checagem_pat_url']
        try:
            print("Realizando requisição para checagem status do PAT Online/Offline")
            response = requests.get(url, verify=False, headers=headers)
            if response.status_code != 200:
                print(f"Erro na requisição. Código de status: {response.status_code, response.text}")
                print(f"API Offline - Data e hora: {current_time}")
                
                # Iniciar o navegador
                options = webdriver.ChromeOptions()
                for option in config['SELENIUM_SETTINGS']['chrome_options'].split(','):
                    options.add_argument(option)
                
                service = Service(executable_path=chromedriver_autoinstaller.install())
                driver = webdriver.Chrome(options=options, service=service)
                driver.maximize_window()
                driver.get("https://atendimento.inss.gov.br/")
                
                tempo_total = 5
                with tqdm(total=100, desc="Aguardando 10 seg para Estabelecimento do PAT..", unit="%", position=0) as pbar:
                    for i in range(101):
                        pbar.update(1)
                        time.sleep(tempo_total / 100)

                print("Aguardando conclusão do procedimento de Login")
                wait = WebDriverWait(driver, 120)
                wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/div/header/div[1]/span")))
                
                for i in range(10, 0, -1):
                    print(f"Aguardando: {i} segundos")
                    time.sleep(1)
                print("Tempo encerrado!")            

                js_script = "return localStorage.getItem('ifs_auth');"
                js_script2 = "return localStorage.getItem('srv_auth');"
                resultado = driver.execute_script(js_script)
                token = driver.execute_script(js_script2)
                
                if not token:
                    raise Exception("Não foi possível obter o token de acesso.")    
                
                if not resultado:
                    raise Exception("Não foi possível obter o token de acesso.")       
                
                headers = {
                    'Authorization': 'Bearer ' + resultado,
                    'Content-type': 'application/json',
                    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36',
                    'tokenservidor': token,
                    'Accept': 'application/json'
                }

                driver.quit()
                with open(headers_file_path, 'w') as file:
                    file.write(json.dumps(headers))

            else:
                pass
                #print(f"API Online - Data e hora: {current_time}")

        except Exception as e:
            print(f"Erro de conexão: {e}")



    obter_token()

    config = configparser.ConfigParser()
    config.read('config.ini')
    CPF = config['API_URLS']['cpf']
    url_cadastro_requerimento = "https://atendimento.inss.gov.br/apis/requerimentosPortalApi/requerimento/cadastro/requerimento"

    # Carrega novamente os headers do arquivo que foram atualizados por 'obter_token'
    headers_file_path = os.path.join(os.getcwd(), config['FILES']['headers_file_path'])
    try:
        with open(headers_file_path, 'r') as file:
            headers = json.loads(file.read())
        print("Headers para requisição POST carregados com sucesso.")
    except Exception as e:
        print(f"Erro ao carregar headers: {e}")
        headers = {}


    print("Iniciando automação...")
    print("Por favor, faça login quando o navegador abrir, caso não tenha feito...")

    # Carregar o arquivo Excel
    arquivo_excel = "protocolo_pat.xlsx"
    workbook = openpyxl.load_workbook(arquivo_excel)
    worksheet = workbook.active  # ou workbook["NomeDaAba"] se quiser uma aba específica


    def obter_servidor_responsavel(cod_servico, headers):
        """
        Faz uma requisição GET ao endpoint de 'responsavel/inicial' para obter
        o ID do servidor responsável. Retorna o 'id' numérico.
        """
        url_responsavel = (
            "https://atendimento.inss.gov.br/"
            f"apis/requerimentosPortalApi/requerimento/cadastro/servico/{cod_servico}/responsavel/inicial"
        )
        resp = requests.get(url_responsavel, headers=headers, verify=False)
        if resp.status_code == 200:
            lista_resp = resp.json()  # geralmente vem um array
            if isinstance(lista_resp, list) and len(lista_resp) > 0:
                return lista_resp[0].get("id")  # por exemplo, 11509
            else:
                raise Exception("Não foi possível encontrar 'id' do responsável no JSON retornado.")
        else:
            raise Exception(
                f"Erro ao obter responsável (status: {resp.status_code}). "
                f"Retorno: {resp.text}"
            )
            
            
    def obter_local_atendimento(cod_servico, cidadao, headers):

        url_local = (
            "https://atendimento.inss.gov.br/"
            f"apis/requerimentosPortalApi/requerimento/cadastro/servico/{cod_servico}/cidadao/{cidadao}/local/atendimento?void"
        )
        resp = requests.get(url_local, headers=headers, verify=False)
        if resp.status_code == 200:
            json_local = resp.json()        
            id_local = json_local["local"]["id"]
            data_disponibilidade_original = json_local["vaga"]["data"]
            dt_local = datetime.fromisoformat(data_disponibilidade_original)
            dt_utc = dt_local.astimezone(timezone.utc)        
            data_disponibilidade = dt_utc.isoformat(timespec="milliseconds").replace("+00:00", "Z")

            return id_local, data_disponibilidade
        else:
            raise Exception(
                f"Erro ao obter local de atendimento (status: {resp.status_code}). "
                f"Retorno: {resp.text}"
            )

    def enviar_comentario(headers, protocolo, despacho):
        """
        Envia um comentário (despacho) para o protocolo informado,
        no endpoint /comentariosApi/comentarios.
        Retorna True (sucesso), False (falha) ou None em caso de exceção.
        """
        url_comentario = "https://atendimento.inss.gov.br/apis/comentariosApi/comentarios"
        data_comentario = {
            "conteudo": despacho,                   # texto do comentário
            "tarefa": {"protocolo": protocolo},     # valor obtido no response anterior
            "sistemaParceiro": {"sistema": "PAT"},
            "canal": "MODULO_TAREFAS",
            "origem": "COMENTARIO",
        }
        try:
            resp = requests.post(url_comentario, json=data_comentario, headers=headers, verify=False)
            if resp.status_code == 201:
                print("Comentário enviado com sucesso.")
                return True
            else:
                print("Erro no envio do comentário.")
                print("Status Code:", resp.status_code)
                print("Resposta:", resp.text)
                return False
        except requests.exceptions.RequestException as e:
            print("Exceção ao enviar comentário:", e)
            return None


    linha = 2
    max_tentativas = 10

    # ------------------------------------------------------------------------------
    # INÍCIO DO SCRIPT PRINCIPAL
    # ------------------------------------------------------------------------------
    while True:
        print("=================================================")
        print(f"Executando linha:{linha}")
        siape = worksheet.cell(row=linha, column=1).value
        cod = worksheet.cell(row=linha, column=2).value    
        despacho = worksheet.cell(row=linha, column=3).value        
        status_atual = worksheet.cell(row=linha, column=4).value
        
        if cod is None:
            break
        
        if status_atual is not None:
            linha +=1        
            continue
        # Calculando a data de disponibilidade (data de hoje + 5 dias), no formato ISO
        data_disponibilidade = (datetime.now() + timedelta(days=5)).strftime("%Y-%m-%dT03:00:00.000Z")
        
        # 1) Obter servidor responsável
        try:
            responsavel_id = obter_servidor_responsavel(cod, headers)
            print(f"ID do Responsável obtido: {responsavel_id}")
        except Exception as err:
            print("[ERRO] Falha ao obter servidoresResponsaveis:", err)
            input("Pressione Enter para sair...")
            sys.exit(1)

        # 2) Obter local de atendimento e dataDisponibilidade
        try:
            id_local_atendimento, data_disponibilidade = obter_local_atendimento(cod, CPF, headers)
            print(f"idLocalAtendimento obtido: {id_local_atendimento}")
            print(f"dataDisponibilidade obtida: {data_disponibilidade}")
        except Exception as err:
            print("[ERRO] Falha ao obter idLocalAtendimento:", err)
            input("Pressione Enter para sair...")
            sys.exit(1)
        
        # Monta o payload    
        payload = {
            "idServico": int(cod) if pd.notnull(cod) else None,
            "origemSolicitacao": "INTRANET",
            "servidoresResponsaveis": [{"id": responsavel_id}],
            "idLocalAtendimento": id_local_atendimento,
            "dataDisponibilidade": data_disponibilidade,
            "requerente": {},
            "houveAtualizacaoDadosCadastrais": False,
            "camposAdicionais": [],
            "anexos": [],
            "vaga": {
                "dataDisponivel": ""
            }
        }
        
        tentativas = 0
        url = "https://atendimento.inss.gov.br/apis/requerimentosPortalApi/requerimento/cadastro/requerimento"

        while True:
            tentativas += 1
            try:
                response = requests.post(url, headers=headers, json=payload, verify=False)           
                print(f"[POST] Tentativa {tentativas} - Status Code: {response.status_code}")
                if response.status_code == 200:
                    print("[OK] Tarefa executada com sucesso.")
                    try:
                        resp_json = response.json()
                        protocolo_req = resp_json["answer"]["protocoloRequerimento"]
                        worksheet.cell(row=linha, column=4).value = protocolo_req
                        print("Protocolo gerado:", protocolo_req)
                        if despacho and isinstance(despacho, str) and despacho.strip():
                            ok_comentario = enviar_comentario(headers, protocolo_req, despacho)
                            if ok_comentario is True:
                                worksheet.cell(row=linha, column=5).value = "Despacho enviado."
                            elif ok_comentario is False:
                                worksheet.cell(row=linha, column=5).value = "Erro no envio do Despacho."
                            else:
                                worksheet.cell(row=linha, column=5).value = "Exceção no envio do Despacho."
                        else:
                            print("Não há despacho para enviar como Despacho.")
                    except Exception as e:
                        print(f"Erro ao extrair protocolo e enviar Despacho: {e}")
                        worksheet.cell(row=linha, column=5).value = f"Exceção: {str(e)}"                    
                    break
                elif response.status_code in [403, 406]:               
                    if tentativas < max_tentativas:
                        print(f"[ERRO] Status {response.status_code}. Tentativa {tentativas} de {max_tentativas}. Aguardando 30 segundos...")
                        time.sleep(30)                    
                        continue
                    else:                    
                        raise Exception(f"Excesso de tentativas (status: {response.status_code}).")
                else:                
                    print(f"[ERRO] Falha não relacionada a 403 ou 406. Status code: {response.status_code}")
                    break
            except Exception as e:            
                worksheet.cell(row=linha, column=4).value = f"Exceção: {str(e)}"
                print("[ERRO] Exceção ao enviar requisição:", e)           
                break
        workbook.save(arquivo_excel)
        linha += 1

    workbook.save(arquivo_excel)
    print("Processamento finalizado, planilha salva.")
    print("Processo finalizado.")

    pass

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("[FATAL] Ocorreu um erro não tratado:")
        print(e)
        input("Pressione Enter para sair...")